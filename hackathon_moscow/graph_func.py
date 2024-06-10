import openpyxl
# cleaner
import os

clear = lambda: os.system('cls')
clear()

# code
wb = openpyxl.load_workbook(filename = '../hackathon_moscow/ГрафДанные.xlsx')
wb_integr_velocity = openpyxl.load_workbook(filename='IntegrVelocity.xlsx')

# active sheet
points = wb['points']
sheet_longitude= wb_integr_velocity['lon']
sheet_lat= wb_integr_velocity['lat']    

# необходимо сделать фукнцию, которая бы считывала (point_id, latitude, longitude, point_name)
# брала в переменные отдельные latitude, longitude - найти эти переменные в IntegrVelocity (пример: A1)
# тогда необходимо найти тож самое в IntegrVelocity.sheet = неделя

def find_coordinate(point_id):
  for row in points.iter_rows(min_row=2):
    if row[0].value == point_id:
      name = row[3].value
      latitude = row[1].value # широта 
      longitude = int(row[2].value) # lon там нет чисел float в табл IntegrVelocity = долгота
      longitude_coordinate = None
      latitude_coordinate = None
      
      # поиск похожей широты в IntegrVelocity
      min_difference = float('inf')
      for col in sheet_lat.iter_cols():
        for cell in col:
          if cell.value is not None:  # проверяем, что ячейка не пустая
            difference = abs(cell.value - latitude)
            if difference < min_difference:
              min_difference = difference
              latitude_coordinate = cell.coordinate
            
      # ищем ячейку долготы
      for col in sheet_longitude.iter_cols():
        for cell in col:
          if cell.value == longitude:
            longitude_coordinate = cell.coordinate
            
      return name, latitude, longitude,  latitude_coordinate, longitude_coordinate #если нет числа
  return None

# 0	73.1	80
# 1	69.4	86.15
# 2	69.9	44.6
# 3	69.15	57.68

point_id = int(input("Input ID:"))
print(find_coordinate(point_id))

